"""
billing_core.py -- GUI-agnostic business logic extracted from the original Tkinter app.

Everything here is pure Python + reportlab/openpyxl/Pillow: GST math, invoice numbering,
PDF generation, Excel generation, and JSON persistence for profiles/invoices. No tkinter,
no Kivy -- so this same module is imported unchanged by:
  - the Windows build (Kivy desktop / PyInstaller), and
  - the Android build (Kivy / Buildozer APK).

Call set_base_dir(path) once at app startup to point all reads/writes at the right folder
for the platform: next to the .exe on Windows, or the folder the user picked (internal vs
external storage) on Android. If you never call it, it defaults to the folder this file
lives in, same as the original desktop app.
"""
import os
import io
import re
import sys
import json
import shutil
import traceback
from datetime import datetime

# --- ReportLab PDF Dependencies ---
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.utils import ImageReader

# --- Pillow (used to fade the logo into a translucent PDF watermark) ---
from PIL import Image as PILImage

# --- openpyxl Dependencies (Excel export) ---
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Persistence Layer for Profiles & Invoices ---
def get_app_base_dir():
    """Anchor every file the app writes (profiles, invoices, counters, logos, generated
    PDFs/Excels) to the folder the program actually lives in -- NOT to the process's
    current working directory.

    CWD is unreliable for a double-clicked/elevated GUI app: Windows can launch it with
    CWD set to C:\\Windows\\System32 -- a protected, admin-only folder -- depending on how
    it's started (a desktop shortcut with no 'Start in' folder set, 'Run as administrator',
    Task Scheduler, etc.), even though the script/exe itself sits somewhere completely
    normal like Desktop or Documents. That mismatch is what caused 'Windows denied access
    while creating the folder ... inside C:\\WINDOWS\\System32' -- the code was writing to
    os.getcwd() instead of to where the program lives."""
    if getattr(sys, "frozen", False):
        # Running as a bundled .exe (e.g. PyInstaller)
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))

APP_BASE_DIR = get_app_base_dir()

PROFILES_FILE = os.path.join(APP_BASE_DIR, "profiles.json")
INVOICES_DB_FILE = os.path.join(APP_BASE_DIR, "invoices.json")
LOGOS_DIR = os.path.join(APP_BASE_DIR, "logos")

def set_base_dir(path):
    """Point every file this module reads/writes (profiles, invoices, counters, logos,
    generated PDFs/Excels) at `path` instead of the default next-to-the-program folder.
    Used on Android to switch between internal (app-private) and external (shared/visible
    in Files app) storage based on what the user picks; on Windows you can call it too if
    you want the data folder somewhere other than next to the .exe (e.g. Documents)."""
    global APP_BASE_DIR, PROFILES_FILE, INVOICES_DB_FILE, LOGOS_DIR
    os.makedirs(path, exist_ok=True)
    APP_BASE_DIR = path
    PROFILES_FILE = os.path.join(APP_BASE_DIR, "profiles.json")
    INVOICES_DB_FILE = os.path.join(APP_BASE_DIR, "invoices.json")
    LOGOS_DIR = os.path.join(APP_BASE_DIR, "logos")
    os.makedirs(LOGOS_DIR, exist_ok=True)

DEFAULT_COMPANY = {
    "name": "GUHAN ENTERPRISES",
    "address": "#49, 8TH Phase, Gottigere Main Road, JP Nagar, Bangalore-560078",
    "gstin": "29QVCPS7350N1ZE",
    "contact": "8867567297",
    "email": "",
    "bank_name": "State Bank of India",
    "ac_no": "421900234511",
    "ifsc": "SBIN0007421",
}

STANDARD_UNITS = ["Nos", "Kgs", "Sets", "Pcs", "Box", "Ltr", "Mtr", "Pair", "Dozen", "Roll", "Bag"]

def load_profiles():
    if os.path.exists(PROFILES_FILE):
        try:
            with open(PROFILES_FILE, "r") as f:
                data = json.load(f)
                data.setdefault("companies", [])
                data.setdefault("buyers", [])
                return data
        except (ValueError, json.JSONDecodeError):
            pass
    return {"companies": [DEFAULT_COMPANY.copy()], "buyers": []}

def save_profiles(profiles):
    with open(PROFILES_FILE, "w") as f:
        json.dump(profiles, f, indent=2)

def load_invoices_db():
    if os.path.exists(INVOICES_DB_FILE):
        try:
            with open(INVOICES_DB_FILE, "r") as f:
                return json.load(f)
        except (ValueError, json.JSONDecodeError):
            pass
    return {}

def save_invoice_to_db(invoice_id, invoice_data):
    db = load_invoices_db()
    db[invoice_id] = invoice_data
    with open(INVOICES_DB_FILE, "w") as f:
        json.dump(db, f, indent=2)

def upsert_profile(profile_list, key_field, new_profile):
    key_val = new_profile.get(key_field, "").strip().lower()
    for i, p in enumerate(profile_list):
        if p.get(key_field, "").strip().lower() == key_val:
            profile_list[i] = new_profile
            return
    profile_list.append(new_profile)

def slugify(text):
    text = text.strip().upper()
    text = re.sub(r"[^A-Z0-9]+", "_", text)
    return text.strip("_") or "DEFAULT"

def ensure_company_folder(company_name):
    """Create (or reuse) the per-company output folder next to the program itself.
    Raises a clear, actionable PermissionError instead of a bare WinError if Windows
    blocks the folder creation -- the most common causes being Controlled Folder Access /
    antivirus, or the program folder being a protected/read-only location."""
    folder = os.path.join(APP_BASE_DIR, slugify(company_name or ""))
    try:
        os.makedirs(folder, exist_ok=True)
    except PermissionError as e:
        raise PermissionError(
            f"Windows denied access while creating the folder:\n{folder}\n\n"
            "This is almost always one of these three causes:\n"
            "1) Windows Security > Virus & threat protection > Manage ransomware protection > "
            "Controlled folder access is ON and is blocking this app. Either add python.exe "
            "(or this program's .exe) to 'Allow an app through Controlled folder access', or "
            "turn Controlled folder access off.\n"
            "2) The program folder is in a protected/read-only location (Program Files, "
            "C:\\Windows, a paused OneDrive folder, a locked network drive). Move the program "
            "to a normal folder such as Documents or Desktop and run it from there.\n"
            "3) A folder or file named '" + os.path.basename(folder) + "' already exists here with "
            "restricted permissions from an earlier run (e.g. it was created while running as a "
            "different Windows user, or as Administrator). Try renaming or deleting it, then run again.\n\n"
            f"Original error: {e}"
        ) from e
    return folder

def sanitize_filename_component(text):
    """Strip characters Windows (and other OSes) forbid in file names. The Document No./
    Invoice No. field is user-editable, so a manually typed value containing e.g. ':' or '?'
    would otherwise silently produce a path that can never be saved. Replacing only '/' (as
    before) missed \\ : * ? " < > | and trailing dots/spaces, which are also invalid on
    Windows."""
    text = (text or "").strip()
    text = re.sub(r'[\\/:*?"<>|]', "_", text)
    text = text.strip(" .")
    return text or "DOC"

def safe_file_save(save_fn, output_path):
    """Run save_fn() (a PDF doc.build(...) or an openpyxl wb.save(...) call) and turn a
    locked-file/permission failure into a clear, actionable error instead of a bare
    PermissionError/OSError. The most common real-world cause is that the PDF or Excel
    file is already open in another program (a viewer, Excel, a browser tab) when the
    app tries to regenerate it."""
    try:
        save_fn()
    except PermissionError as e:
        raise PermissionError(
            f"Could not save:\n{os.path.abspath(output_path)}\n\n"
            "This almost always means one of the following:\n"
            "1) The file is currently open in another program (a PDF viewer, Excel, or a "
            "browser tab showing it). Close it there and try saving again.\n"
            "2) Windows Controlled Folder Access is blocking this app from writing here "
            "(see the folder-creation error message for how to allow it).\n"
            "3) You don't have write permission to this folder.\n\n"
            f"Original error: {e}"
        ) from e
    except OSError as e:
        raise OSError(
            f"Could not save:\n{os.path.abspath(output_path)}\n\n"
            "This usually means the file name contains a character your operating system "
            "doesn't allow -- most often from a manually typed Document No./Invoice No. "
            "field -- or the full path is too long.\n\n"
            f"Original error: {e}"
        ) from e
    return output_path

# --- Dynamic Automated Counters ---
DOC_PREFIX_MAP = {
    "Tax Invoice": ("INV", "counter_invoice"),
    "Proforma Invoice": ("PRO", "counter_proforma"),
    "Quotation": ("QTN", "counter_quotation"),
}

def _read_local_counter(file_path):
    """Return the last number saved locally for this counter file, or None if never issued."""
    if os.path.exists(file_path):
        with open(file_path, "r") as f:
            try:
                return int(f.read().strip())
            except ValueError:
                return None
    return None

def get_next_document_number(doc_type, company_name, cloud_client=None):
    """
    Work out the next Tax Invoice / Proforma / Quotation number.

    Behaviour:
    - Offline, or no cloud_client passed in: identical to the original app -- reads the
      local counter_*.txt file next to the data folder and increments it.
    - Online with a cloud_client (see cloud_sync.py): asks the shared cloud counter for the
      next number so two devices (e.g. the Windows PC and a phone) don't hand out the same
      invoice number, then keeps the local file in sync as a fallback for the next
      offline session.

    cloud_client must implement: reserve_next(company_slug, counter_key, local_last) -> int
    and should raise cloud_sync.CloudUnavailable if it can't reach the server (network down,
    timeout, etc.) -- that's caught here and we transparently fall back to local numbering.
    """
    current_year = datetime.now().year
    next_year_short = str(current_year + 1)[-2:]
    year_format = f"{current_year}-{next_year_short}"

    doc_tag, file_base = DOC_PREFIX_MAP.get(doc_type, ("INV", "counter_invoice"))
    company_slug = slugify(company_name or "DEFAULT")
    file_path = os.path.join(APP_BASE_DIR, f"{file_base}_{company_slug}.txt")

    initials = "".join(w[0] for w in re.findall(r"[A-Za-z0-9]+", company_name or "")) or "GE"
    prefix = f"{initials.upper()}-{doc_tag}"

    local_last = _read_local_counter(file_path)

    next_num = None
    if cloud_client is not None:
        try:
            next_num = cloud_client.reserve_next(company_slug, file_base, local_last)
        except Exception:
            next_num = None  # CloudUnavailable or any other network/service failure

    if next_num is None:
        # Offline path (or no cloud_client at all) -- exactly the original local behaviour.
        next_num = (local_last + 1) if local_last is not None else 1023

    return f"{prefix}-{next_num}/{year_format}", next_num, file_path

def save_document_number(file_path, num):
    with open(file_path, "w") as f:
        f.write(str(num))

def num_to_words(number):
    number = int(round(number))
    if number == 0:
        return "Zero Rupees Only"
    units = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine", "Ten", 
             "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
    
    def convert_chunk(n):
        out = ""
        if n >= 100:
            out += units[n // 100] + " Hundred"
            n %= 100
            out += " and " if n > 0 else " "
        if n >= 20:
            out += tens[n // 10] + " "
            n %= 10
        if n > 0:
            out += units[n] + " "
        return out.strip()

    res = ""
    if number >= 10000000:
        res += convert_chunk(number // 10000000) + " Crore "
        number %= 10000000
    if number >= 100000:
        res += convert_chunk(number // 100000) + " Lakh "
        number %= 100000
    if number >= 1000:
        res += convert_chunk(number // 1000) + " Thousand "
        number %= 1000
    if number > 0:
        res += convert_chunk(number)
        
    return res.strip() + " Rupees Only"

def fmt_rate(rate):
    """Format a percentage rate without redundant trailing zeros (18.0 -> '18', 12.5 -> '12.5')."""
    r = round(float(rate), 2)
    if r == int(r):
        return str(int(r))
    return f"{r:g}"

def compute_item_tax(item):
    """Return the taxable amount and tax split (SGST/CGST or IGST) for a single line item."""
    amount = item["Quantity"] * item["Unit Price"]
    gst_type = item.get("GST Type", "GST")
    # Default to 18% for older saved items created before per-item GST/IGST selection existed.
    rate = float(item.get("GST Rate", 18) if item.get("GST Rate") is not None else 18)

    if gst_type == "IGST":
        igst = amount * rate / 100.0
        return {"amount": amount, "gst_type": "IGST", "rate": rate, "sgst": 0.0, "cgst": 0.0, "igst": igst, "tax_total": igst}
    else:
        half = rate / 2.0
        sgst = amount * half / 100.0
        cgst = amount * half / 100.0
        return {"amount": amount, "gst_type": "GST", "rate": rate, "sgst": sgst, "cgst": cgst, "igst": 0.0, "tax_total": sgst + cgst}

def group_items_by_gst(items_list):
    """
    Group line items by identical (GST Type, Rate) so the invoice can show one consolidated
    tax line per rate/type combination, listing which SI numbers fall under it (e.g. "1, 2").
    Returns a list of group dicts in first-seen order.
    """
    groups = {}
    order = []
    for item in items_list:
        tax = compute_item_tax(item)
        key = (tax["gst_type"], round(tax["rate"], 2))
        if key not in groups:
            groups[key] = {
                "gst_type": tax["gst_type"], "rate": tax["rate"],
                "si_nos": [], "taxable": 0.0, "sgst": 0.0, "cgst": 0.0, "igst": 0.0, "tax_total": 0.0,
            }
            order.append(key)
        g = groups[key]
        g["si_nos"].append(item["SI No."])
        g["taxable"] += tax["amount"]
        g["sgst"] += tax["sgst"]
        g["cgst"] += tax["cgst"]
        g["igst"] += tax["igst"]
        g["tax_total"] += tax["tax_total"]
    return [groups[key] for key in order]

def group_si_label(si_nos):
    """Turn a list of SI numbers into a compact label like '1, 2, 5'."""
    return ", ".join(str(n) for n in sorted(si_nos))

def round_grand_total(raw_total):
    """
    Round the grand total to the nearest rupee using standard paise rounding:
    0 to 49 paise rounds DOWN (subtract), 50 to 99 paise rounds UP (add).
    Returns (grand_total, round_off) where round_off = grand_total - raw_total.
    """
    raw_total_r = round(raw_total, 2)
    rupees = int(raw_total_r)
    paise = round((raw_total_r - rupees) * 100)
    if paise >= 50:
        grand_total = rupees + 1
    else:
        grand_total = rupees
    round_off = grand_total - raw_total_r
    return grand_total, round_off

def get_scaled_logo_flowable(logo_path, max_width=90, max_height=68):
    """Build a ReportLab Image flowable for the company logo, scaled to fit within
    max_width x max_height while preserving aspect ratio. Returns None if there is
    no logo configured or the file can't be read."""
    if not logo_path or not os.path.exists(logo_path):
        return None
    try:
        reader = ImageReader(logo_path)
        iw, ih = reader.getSize()
        if iw <= 0 or ih <= 0:
            return None
        scale = min(max_width / iw, max_height / ih)
        return Image(logo_path, width=iw * scale, height=ih * scale)
    except Exception:
        return None

def build_watermark_reader(logo_path, opacity=0.10):
    """Build a faded, semi-transparent copy of the logo (as an in-memory PNG) for use
    as a page watermark. Returns (ImageReader, (width, height)) or None if there's no
    logo configured or the file can't be read."""
    if not logo_path or not os.path.exists(logo_path):
        return None
    try:
        img = PILImage.open(logo_path).convert("RGBA")
        r, g, b, a = img.split()
        # Scale the alpha channel down so the logo prints as a faint background mark
        # rather than a solid image sitting on top of the invoice content.
        a = a.point(lambda v: int(v * opacity))
        faded = PILImage.merge("RGBA", (r, g, b, a))
        buf = io.BytesIO()
        faded.save(buf, format="PNG")
        buf.seek(0)
        return ImageReader(buf), img.size
    except Exception:
        return None

def draw_page_border(canvas, doc, logo_path=""):
    canvas.saveState()

    # Watermark first so it sits behind the border/content drawn on top of it.
    watermark = build_watermark_reader(logo_path)
    if watermark:
        reader, (iw, ih) = watermark
        max_w, max_h = A4[0] * 0.55, A4[1] * 0.55
        scale = min(max_w / iw, max_h / ih)
        w, h = iw * scale, ih * scale
        x, y = (A4[0] - w) / 2, (A4[1] - h) / 2
        canvas.drawImage(reader, x, y, width=w, height=h, mask='auto')

    canvas.setStrokeColor(colors.HexColor('#000000'))
    canvas.setLineWidth(1)
    canvas.rect(20, 20, A4[0] - 40, A4[1] - 40)
    canvas.restoreState()

def generate_pdf(doc_type, company_info, buyer_info, bank_info, terms, meta_info, items_list):
    clean_inv_name = sanitize_filename_component(meta_info['invoice_no'])
    file_prefix = doc_type.replace(" ", "_")
    company_folder = ensure_company_folder(company_info.get('name', ''))
    pdf_file = os.path.join(company_folder, f"{file_prefix}_{clean_inv_name}.pdf")

    PAGE_LEFT_MARGIN = 40
    PAGE_RIGHT_MARGIN = 40
    PAGE_TOP_MARGIN = 40
    content_width = A4[0] - PAGE_LEFT_MARGIN - PAGE_RIGHT_MARGIN

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=16, leading=20, textColor=colors.HexColor('#1F4E78'))
    subtitle_style = ParagraphStyle('Sub', parent=styles['Normal'], fontName='Helvetica', fontSize=10, leading=14, textColor=colors.HexColor('#4A4A4A'))
    right_meta_style = ParagraphStyle('Meta', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=12, leading=16, alignment=2)
    right_sub_style = ParagraphStyle('MetaSub', parent=styles['Normal'], fontName='Helvetica', fontSize=10, leading=14, alignment=2)
    table_header_style = ParagraphStyle('TH', fontName='Helvetica-Bold', fontSize=10, leading=12, textColor=colors.white, alignment=1)
    table_cell_style = ParagraphStyle('TC', fontName='Helvetica', fontSize=10, leading=13)
    table_cell_center = ParagraphStyle('TCC', fontName='Helvetica', fontSize=10, leading=13, alignment=1)
    table_cell_right = ParagraphStyle('TCR', fontName='Helvetica', fontSize=10, leading=13, alignment=2)
    normal_footer_style = ParagraphStyle('NF', fontName='Helvetica', fontSize=9, leading=12)

    # --- Bank details, terms & conditions, and signature now flow as normal content that
    # comes right after the item ledger/summary block, instead of being pinned to a fixed
    # position at the bottom of the page. If the ledger runs long, this block simply moves
    # down (or onto the next page) with it instead of overlapping it.
    bank_html = f"<b>Company's Bank Details:</b><br/>Bank Name: {bank_info['bank_name']}<br/>A/c No: {bank_info['ac_no']}<br/>IFSC: {bank_info['ifsc']}"
    terms_html = f"<b>Terms and Conditions:</b><br/>" + "<br/>".join([f"{i+1}. {t}" for i, t in enumerate(terms)])
    footer_left_flow = [Paragraph(bank_html, normal_footer_style), Spacer(1, 5), Paragraph(terms_html, normal_footer_style)]
    footer_right_flow = [Paragraph(f"For {company_info['name'].upper()}", ParagraphStyle('RFT', fontName='Helvetica-Bold', fontSize=10, alignment=2)), Spacer(1, 45), Paragraph("Authorized Signatory", ParagraphStyle('RFB', fontName='Helvetica', fontSize=10, alignment=2))]
    footer_table = Table([[footer_left_flow, footer_right_flow]], colWidths=[300, 215])
    footer_table.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP'), ('BOX', (0,0), (0,0), 0.5, colors.HexColor('#E0E0E0')), ('PADDING', (0,0), (-1,-1), 6)]))

    PAGE_BOTTOM_MARGIN = 40
    doc = SimpleDocTemplate(pdf_file, pagesize=A4, rightMargin=PAGE_RIGHT_MARGIN, leftMargin=PAGE_LEFT_MARGIN, topMargin=PAGE_TOP_MARGIN, bottomMargin=PAGE_BOTTOM_MARGIN)
    story = []

    company_addr_line = f"Address: {company_info['address']}<br/>GSTIN: {company_info['gstin']} | Contact: {company_info['contact']}"
    if company_info.get('email'):
        company_addr_line += f" | Email: {company_info['email']}"
    doc_meta_lines = [f"Document No: {meta_info['invoice_no']}", f"Date: {meta_info['invoice_date']}"]
    if meta_info.get('po_number'):
        doc_meta_lines.append(f"PO No: {meta_info['po_number']}")
    if meta_info.get('po_date'):
        doc_meta_lines.append(f"PO Date: {meta_info['po_date']}")
    logo_flowable = get_scaled_logo_flowable(company_info.get('logo_path', ''))
    if logo_flowable:
        header_data = [
            [logo_flowable, Paragraph(company_info['name'].upper(), title_style), Paragraph(doc_type.upper(), right_meta_style)],
            ["", Paragraph(company_addr_line, subtitle_style), Paragraph("<br/>".join(doc_meta_lines), right_sub_style)]
        ]
        header_table = Table(header_data, colWidths=[95, 220, 200])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'TOP'), ('BOTTOMPADDING', (0,0), (-1,-1), 0),
            ('SPAN', (0,0), (0,1)), ('ALIGN', (0,0), (0,1), 'LEFT'),
        ]))
    else:
        header_data = [
            [Paragraph(company_info['name'].upper(), title_style), Paragraph(doc_type.upper(), right_meta_style)],
            [Paragraph(company_addr_line, subtitle_style), Paragraph("<br/>".join(doc_meta_lines), right_sub_style)]
        ]
        header_table = Table(header_data, colWidths=[315, 200])
        header_table.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP'), ('BOTTOMPADDING', (0,0), (-1,-1), 0)]))
    story.append(header_table)
    story.append(Spacer(1, 10))
    
    buyer_html = f"<b>BUYER / CLIENT: {buyer_info['name'].upper()}</b><br/>{buyer_info['address']}"
    if buyer_info.get('gstin'):
        buyer_html += f"<br/><b>GSTIN:</b> {buyer_info['gstin']}"

    contact_lines = []
    if buyer_info.get('contact_name'):
        contact_lines.append(f"<b>Contact Name:</b> {buyer_info['contact_name']}")
    if buyer_info.get('mobile'):
        contact_lines.append(f"<b>Mobile No:</b> {buyer_info['mobile']}")
    buyer_contact_style = ParagraphStyle('BuyerContact', parent=subtitle_style, alignment=2)
    contact_p = Paragraph("<br/>".join(contact_lines), buyer_contact_style) if contact_lines else Paragraph("", buyer_contact_style)

    buyer_table = Table([[Paragraph(buyer_html, subtitle_style), contact_p]], colWidths=[315, 200])
    buyer_table.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F4F4F4')), ('PADDING', (0,0), (-1,-1), 8), ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#D9D9D9')), ('VALIGN', (0,0), (-1,-1), 'TOP')]))
    story.append(buyer_table)
    story.append(Spacer(1, 10))
    
    LEDGER_COL_WIDTHS = [30, 210, 55, 32, 45, 68, 75]
    header_row = [Paragraph("SI No.", table_header_style), Paragraph("Description of Goods", table_header_style), Paragraph("HSN/SAC", table_header_style), Paragraph("Qty", table_header_style), Paragraph("Unit", table_header_style), Paragraph("Unit Price", table_header_style), Paragraph("Amount", table_header_style)]
    item_rows = []
    subtotal = 0.0
    for item in items_list:
        tax = compute_item_tax(item)
        subtotal += tax["amount"]
        item_rows.append([Paragraph(str(item["SI No."]), table_cell_center), Paragraph(item["Description of Goods"], table_cell_style), Paragraph(item["HSN/SAC"], table_cell_center), Paragraph(str(item["Quantity"]), table_cell_center), Paragraph(item.get("Unit", ""), table_cell_center), Paragraph(f"{item['Unit Price']:.2f}", table_cell_right), Paragraph(f"{tax['amount']:.2f}", table_cell_right)])

    # Consolidate identical GST/IGST rate selections into grouped tax lines, each one
    # naming the SI numbers it covers (e.g. "Items 1, 2").
    tax_groups = group_items_by_gst(items_list)
    total_tax = sum(g["tax_total"] for g in tax_groups)
    raw_total = subtotal + total_tax
    grand_total, round_off = round_grand_total(raw_total)

    # NOTE: the left-hand block (columns 0-4) is SPANned across every summary row below,
    # so reportlab only ever renders the content of its top-left cell. Anything meant to
    # appear in that region -- amount in words AND the GST/IGST group notes -- must
    # therefore be assembled into ONE combined paragraph rather than placed on separate rows.
    gst_note_lines = []
    for g in tax_groups:
        si_label = group_si_label(g["si_nos"])
        if g["gst_type"] == "IGST":
            gst_note_lines.append(f"IGST @{fmt_rate(g['rate'])}% applies to Items: {si_label}")
        else:
            gst_note_lines.append(f"GST @{fmt_rate(g['rate'])}% (SGST {fmt_rate(g['rate']/2)}% + CGST {fmt_rate(g['rate']/2)}%) applies to Items: {si_label}")

    words_html = f"<b>Amount Chargeable (in words):</b><br/>{num_to_words(grand_total)}"
    if gst_note_lines:
        words_html += "<br/><br/><b>GST Break-up:</b><br/>" + "<br/>".join(gst_note_lines)
    words_p = Paragraph(words_html, subtitle_style)

    def lbl_p(txt, bold=False):
        return Paragraph(txt, ParagraphStyle('L', fontName='Helvetica-Bold' if bold else 'Helvetica', fontSize=10, alignment=2))

    summary_rows = [
        [words_p, "", "", "", "", lbl_p("Sub Total:"), lbl_p(f"{subtotal:,.2f}")],
    ]
    for g in tax_groups:
        if g["gst_type"] == "IGST":
            summary_rows.append(["", "", "", "", "", lbl_p(f"IGST ({fmt_rate(g['rate'])}%):"), lbl_p(f"{g['igst']:,.2f}")])
        else:
            summary_rows.append(["", "", "", "", "", lbl_p(f"SGST ({fmt_rate(g['rate']/2)}%):"), lbl_p(f"{g['sgst']:,.2f}")])
            summary_rows.append(["", "", "", "", "", lbl_p(f"CGST ({fmt_rate(g['rate']/2)}%):"), lbl_p(f"{g['cgst']:,.2f}")])
    summary_rows.append(["", "", "", "", "", lbl_p("Round Off:"), lbl_p(f"{round_off:,.2f}")])
    summary_rows.append(["", "", "", "", "", lbl_p("Grand Total:", bold=True), lbl_p(f"{grand_total:,.2f}", bold=True)])

    def build_ledger_style(num_item_rows, num_summary_rows):
        style = [
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E78')), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOTTOMPADDING', (0,0), (-1,0), 6), ('TOPPADDING', (0,0), (-1,0), 6),
            ('GRID', (0,0), (-1, num_item_rows), 0.5, colors.HexColor('#D9D9D9')), ('BOX', (0,0), (-1, -1), 0.5, colors.HexColor('#D9D9D9')),
        ]
        summary_start = num_item_rows + 1
        summary_end = summary_start + num_summary_rows - 1
        style.append(('SPAN', (0, summary_start), (4, summary_end)))
        style.append(('VALIGN', (0, summary_start), (4, summary_end), 'TOP'))
        style.append(('PADDING', (0, summary_start), (4, summary_end), 6))
        style.append(('LINEABOVE', (5, summary_start), (6, -1), 0.5, colors.black))
        style.append(('LINEBELOW', (5, -1), (6, -1), 1.5, colors.black))
        style.append(('GRID', (5, summary_start), (6, -1), 0.5, colors.HexColor('#D9D9D9')))
        # No BACKGROUND fill on the amount-in-words/GST breakup block (cols 0-4) so the
        # page watermark stays visible behind it instead of being painted over.
        return TableStyle(style)

    # The ledger now contains exactly one row per item -- no blank filler rows -- so its
    # height is always driven by how many items were actually entered.
    ledger_data = [header_row] + item_rows + summary_rows
    ledger_table = Table(ledger_data, colWidths=LEDGER_COL_WIDTHS, repeatRows=1)
    ledger_table.setStyle(build_ledger_style(len(item_rows), len(summary_rows)))
    story.append(ledger_table)
    story.append(Spacer(1, 14))

    # Bank details / terms / signature come immediately after the ledger in the normal
    # document flow, so they sit right below however many item rows were entered instead
    # of being anchored to a fixed spot on the page.
    story.append(footer_table)

    logo_path_for_watermark = company_info.get('logo_path', '')
    page_decorator = lambda cnv, d: draw_page_border(cnv, d, logo_path_for_watermark)
    safe_file_save(lambda: doc.build(story, onFirstPage=page_decorator, onLaterPages=page_decorator), pdf_file)
    return pdf_file

# --- Excel Sheet Export Engine ---
XL_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
XL_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF")
XL_TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="1F4E78")
XL_LABEL_FONT = Font(name="Calibri", bold=True, size=10)
XL_NORMAL_FONT = Font(name="Calibri", size=10)
XL_MUTED_FONT = Font(name="Calibri", size=10, color="4A4A4A")
XL_THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)
XL_MONEY_FMT = '#,##0.00'

def generate_excel(doc_type, company_info, buyer_info, bank_info, terms, meta_info, items_list):
    clean_inv_name = sanitize_filename_component(meta_info['invoice_no'])
    file_prefix = doc_type.replace(" ", "_")
    company_folder = ensure_company_folder(company_info.get('name', ''))
    xlsx_file = os.path.join(company_folder, f"{file_prefix}_{clean_inv_name}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = doc_type[:31]

    for col, width in zip("ABCDEFGHIJKLM", [8, 36, 12, 8, 10, 14, 16, 10, 10, 14, 14, 14, 16]):
        ws.column_dimensions[col].width = width

    table_header_row = 1
    headers = ["SI No.", "Description of Goods", "HSN/SAC", "Qty", "Unit", "Unit Price", "Amount",
               "GST Type", "Rate (%)", "SGST", "CGST", "IGST", "Total"]
    for i, htext in enumerate(headers):
        cell = ws.cell(row=table_header_row, column=i + 1, value=htext)
        cell.font = XL_HEADER_FONT
        cell.fill = XL_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = XL_THIN_BORDER

    first_item_row = table_header_row + 1
    for offset, item in enumerate(items_list):
        r = first_item_row + offset
        tax = compute_item_tax(item)
        ws.cell(row=r, column=1, value=item["SI No."]).alignment = Alignment(horizontal="center")
        desc_cell = ws.cell(row=r, column=2, value=item["Description of Goods"])
        desc_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.cell(row=r, column=3, value=item["HSN/SAC"]).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=4, value=item["Quantity"]).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=5, value=item.get("Unit", "")).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=6, value=item["Unit Price"]).number_format = XL_MONEY_FMT
        amt_cell = ws.cell(row=r, column=7, value=f"=D{r}*F{r}")
        amt_cell.number_format = XL_MONEY_FMT
        ws.cell(row=r, column=8, value=tax["gst_type"]).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=9, value=tax["rate"]).alignment = Alignment(horizontal="center")
        sgst_cell = ws.cell(row=r, column=10, value=f'=IF(H{r}="GST",G{r}*(I{r}/2)/100,0)')
        sgst_cell.number_format = XL_MONEY_FMT
        cgst_cell = ws.cell(row=r, column=11, value=f'=IF(H{r}="GST",G{r}*(I{r}/2)/100,0)')
        cgst_cell.number_format = XL_MONEY_FMT
        igst_cell = ws.cell(row=r, column=12, value=f'=IF(H{r}="IGST",G{r}*I{r}/100,0)')
        igst_cell.number_format = XL_MONEY_FMT
        total_cell = ws.cell(row=r, column=13, value=f"=G{r}+J{r}+K{r}+L{r}")
        total_cell.number_format = XL_MONEY_FMT
        for c in range(1, 14):
            cell = ws.cell(row=r, column=c)
            cell.font = XL_NORMAL_FONT
            cell.border = XL_THIN_BORDER

        # Auto-estimate a row height so wrapped descriptions aren't clipped
        desc_col_width = 36  # matches column B width set above
        est_lines = max(1, -(-len(item["Description of Goods"]) // desc_col_width))
        ws.row_dimensions[r].height = max(15, est_lines * 15)

    last_item_row = first_item_row + len(items_list) - 1 if items_list else first_item_row - 1

    # --- Grouped GST/IGST Summary Block ---
    # One row per distinct (GST Type, Rate) combination actually used on the invoice,
    # each naming the SI numbers it covers, followed by Sub Total / Round Off / Grand Total.
    subtotal = sum(item["Quantity"] * item["Unit Price"] for item in items_list)
    tax_groups = group_items_by_gst(items_list)
    total_tax = sum(g["tax_total"] for g in tax_groups)
    raw_total = subtotal + total_tax
    grand_total, round_off = round_grand_total(raw_total)

    r = last_item_row + 2
    ws.cell(row=r, column=1, value="GST Break-up").font = XL_LABEL_FONT
    r += 1
    gh_row = r
    for i, htext in enumerate(["Items (SI No.)", "GST Type", "Rate (%)", "Taxable Value", "SGST", "CGST", "IGST"]):
        cell = ws.cell(row=gh_row, column=i + 1, value=htext)
        cell.font = XL_HEADER_FONT
        cell.fill = XL_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = XL_THIN_BORDER
    r += 1
    for g in tax_groups:
        ws.cell(row=r, column=1, value=group_si_label(g["si_nos"])).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=g["gst_type"]).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3, value=g["rate"]).alignment = Alignment(horizontal="center")
        for col, val in [(4, g["taxable"]), (5, g["sgst"]), (6, g["cgst"]), (7, g["igst"])]:
            c = ws.cell(row=r, column=col, value=val)
            c.number_format = XL_MONEY_FMT
            c.alignment = Alignment(horizontal="right")
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            cell.font = XL_NORMAL_FONT
            cell.border = XL_THIN_BORDER
        r += 1

    r += 1
    totals_panel = [
        ("Sub Total:", subtotal),
        ("Total GST/IGST:", total_tax),
        ("Round Off:", round_off),
        ("Grand Total:", grand_total),
    ]
    for label, val in totals_panel:
        lbl_cell = ws.cell(row=r, column=6, value=label)
        lbl_cell.font = XL_LABEL_FONT
        lbl_cell.alignment = Alignment(horizontal="right")
        val_cell = ws.cell(row=r, column=7, value=val)
        val_cell.font = XL_LABEL_FONT
        val_cell.number_format = XL_MONEY_FMT
        r += 1

    ws.print_area = f"A1:M{r}"

    safe_file_save(lambda: wb.save(xlsx_file), xlsx_file)
    return xlsx_file

# --- Billing Summary Register (auto-updating across every document generated) ---
SUMMARY_FILE_NAME = "Billing_Summary.xlsx"
SUMMARY_HEADERS = ["SI No.", "Document Type", "Bill/Document No.", "Date", "Buyer Name", "Total Amount", "GST", "Grand Total"]

def compute_invoice_totals(items_list):
    subtotal = sum(item["Quantity"] * item["Unit Price"] for item in items_list)
    tax_groups = group_items_by_gst(items_list)
    gst_total = sum(g["tax_total"] for g in tax_groups)
    raw_total = subtotal + gst_total
    grand_total, _round_off = round_grand_total(raw_total)
    return subtotal, gst_total, grand_total

def _find_last_summary_row(ws):
    r = 2
    while ws.cell(row=r, column=3).value not in (None, ""):
        r += 1
    return r - 1  # last populated data row; returns 1 (header only) if no data yet

def update_summary_excel(company_info, doc_type, meta_info, buyer_info, items_list):
    company_folder = ensure_company_folder(company_info.get('name', ''))
    summary_path = os.path.join(company_folder, SUMMARY_FILE_NAME)

    subtotal, gst_total, grand_total = compute_invoice_totals(items_list)
    bill_no = meta_info['invoice_no']
    bill_date = meta_info.get('invoice_date', '')
    buyer_name = buyer_info.get('name', '')

    if os.path.exists(summary_path):
        wb = load_workbook(summary_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        for col, width in zip("ABCDEFGH", [8, 16, 22, 14, 30, 16, 16, 16]):
            ws.column_dimensions[col].width = width
        for i, htext in enumerate(SUMMARY_HEADERS):
            cell = ws.cell(row=1, column=i + 1, value=htext)
            cell.font = XL_HEADER_FONT
            cell.fill = XL_HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = XL_THIN_BORDER
        ws.freeze_panes = "A2"

        # Live summary panel off to the side -- formulas cover a generous fixed range so
        # they keep including new rows automatically as more documents are added, with no
        # need to move or rewrite them on every update.
        ws.column_dimensions["I"].width = 3
        ws.column_dimensions["J"].width = 20
        ws.column_dimensions["K"].width = 16
        panel = [
            ("Total Documents:", "=COUNTA(C2:C5000)", None),
            ("Total Amount (Sum):", "=SUM(F2:F5000)", XL_MONEY_FMT),
            ("Total GST (Sum):", "=SUM(G2:G5000)", XL_MONEY_FMT),
            ("Grand Total (Sum):", "=SUM(H2:H5000)", XL_MONEY_FMT),
        ]
        for offset, (label, formula, fmt) in enumerate(panel):
            r = 2 + offset
            lbl_cell = ws.cell(row=r, column=10, value=label)
            lbl_cell.font = XL_LABEL_FONT
            val_cell = ws.cell(row=r, column=11, value=formula)
            val_cell.font = XL_LABEL_FONT
            if fmt:
                val_cell.number_format = fmt

    last_data_row = _find_last_summary_row(ws)

    target_row = None
    for r in range(2, last_data_row + 1):
        if ws.cell(row=r, column=3).value == bill_no:
            target_row = r
            break

    if target_row is None:
        target_row = last_data_row + 1
        si_no = target_row - 1
    else:
        si_no = ws.cell(row=target_row, column=1).value or (target_row - 1)

    row_values = [si_no, doc_type, bill_no, bill_date, buyer_name, subtotal, gst_total, grand_total]
    for i, val in enumerate(row_values):
        cell = ws.cell(row=target_row, column=i + 1, value=val)
        cell.font = XL_NORMAL_FONT
        cell.border = XL_THIN_BORDER
        if i in (5, 6, 7):
            cell.number_format = XL_MONEY_FMT
            cell.alignment = Alignment(horizontal="right")
        else:
            cell.alignment = Alignment(horizontal="center" if i != 4 else "left")

    ws.auto_filter.ref = f"A1:H{target_row}"
    safe_file_save(lambda: wb.save(summary_path), summary_path)
    return summary_path

# --- HSN-wise Summary Register (single consolidated table per quarter -- every bill's ---
# --- items are grouped/merged into one row per HSN code, no bill/buyer/date columns) ---
QUARTER_DEFS = [(1, 3, "Jan-Mar"), (4, 6, "Apr-Jun"), (7, 9, "Jul-Sep"), (10, 12, "Oct-Dec")]
HSN_SUMMARY_HEADERS = ["HSN/SAC", "Total Quantity", "Total Unit Price", "Total CGST", "Total SGST", "Total IGST"]

def get_quarter_file_tag(date_str):
    """Map an invoice date (dd-mm-yyyy) to its calendar quarter, e.g. ('Jan-Mar_2026', 'Jan-Mar', 2026).
    Falls back to today's date if the string is missing or unparsable."""
    try:
        d = datetime.strptime((date_str or "").strip(), "%d-%m-%Y")
    except (ValueError, AttributeError):
        d = datetime.now()
    for start, end, label in QUARTER_DEFS:
        if start <= d.month <= end:
            return f"{label}_{d.year}", label, d.year
    return f"Qtr_{d.year}", "Qtr", d.year  # unreachable safeguard

def group_items_by_hsn(items_list):
    """Group line items by HSN/SAC code. Returns a list of dicts (first-seen order) each with:
    summed quantity, summed taxable value ('Unit Price' total, i.e. Qty x Rate), and summed CGST/SGST/IGST."""
    groups = {}
    order = []
    for item in items_list:
        hsn = (item.get("HSN/SAC") or "").strip() or "N/A"
        tax = compute_item_tax(item)
        if hsn not in groups:
            groups[hsn] = {"hsn": hsn, "qty": 0, "amount": 0.0, "cgst": 0.0, "sgst": 0.0, "igst": 0.0}
            order.append(hsn)
        g = groups[hsn]
        g["qty"] += item.get("Quantity", 0)
        g["amount"] += tax["amount"]
        g["cgst"] += tax["cgst"]
        g["sgst"] += tax["sgst"]
        g["igst"] += tax["igst"]
    return [groups[h] for h in order]

def _read_existing_hsn_totals(ws):
    """Read the current accumulated HSN rows (row 2 onward) back into an order list + totals dict,
    stopping at the first blank row or the existing 'Grand Total' row (which is rebuilt, not re-read)."""
    order = []
    totals = {}
    r = 2
    while True:
        hsn_val = ws.cell(row=r, column=1).value
        if hsn_val in (None, "", "Grand Total"):
            break
        totals[hsn_val] = {
            "qty": ws.cell(row=r, column=2).value or 0,
            "amount": ws.cell(row=r, column=3).value or 0.0,
            "cgst": ws.cell(row=r, column=4).value or 0.0,
            "sgst": ws.cell(row=r, column=5).value or 0.0,
            "igst": ws.cell(row=r, column=6).value or 0.0,
        }
        order.append(hsn_val)
        r += 1
    return order, totals

def _write_hsn_summary_sheet(ws, order, totals):
    """Clear the sheet and rewrite the full HSN table (header + one row per HSN code + Grand Total)."""
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)

    for i, htext in enumerate(HSN_SUMMARY_HEADERS):
        cell = ws.cell(row=1, column=i + 1, value=htext)
        cell.font = XL_HEADER_FONT
        cell.fill = XL_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = XL_THIN_BORDER
    ws.freeze_panes = "A2"

    r = 2
    for hsn in order:
        t = totals[hsn]
        row_vals = [hsn, t["qty"], t["amount"], t["cgst"], t["sgst"], t["igst"]]
        for i, val in enumerate(row_vals):
            cell = ws.cell(row=r, column=i + 1, value=val)
            cell.font = XL_NORMAL_FONT
            cell.border = XL_THIN_BORDER
            if i >= 2:
                cell.number_format = XL_MONEY_FMT
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="center")
        r += 1

    grand_vals = [
        "Grand Total",
        sum(t["qty"] for t in totals.values()),
        sum(t["amount"] for t in totals.values()),
        sum(t["cgst"] for t in totals.values()),
        sum(t["sgst"] for t in totals.values()),
        sum(t["igst"] for t in totals.values()),
    ]
    for i, val in enumerate(grand_vals):
        cell = ws.cell(row=r, column=i + 1, value=val)
        cell.font = XL_LABEL_FONT
        cell.fill = PatternFill("solid", fgColor="F2F2F2")
        cell.border = XL_THIN_BORDER
        if i >= 2:
            cell.number_format = XL_MONEY_FMT
            cell.alignment = Alignment(horizontal="right")
        else:
            cell.alignment = Alignment(horizontal="center")

    widths = [14, 14, 16, 14, 14, 14]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.auto_filter.ref = f"A1:F{r}"

def update_hsn_summary_excel(company_info, meta_info, items_list):
    """Merge this bill's HSN-wise totals into the single consolidated table for the current
    quarter (Jan-Mar/Apr-Jun/Jul-Sep/Oct-Dec). Every bill updates the same HSN rows -- matching
    HSN codes accumulate into one row instead of each bill getting its own block."""
    company_folder = ensure_company_folder(company_info.get('name', ''))

    quarter_tag, quarter_label, _year = get_quarter_file_tag(meta_info.get('invoice_date', ''))
    file_path = os.path.join(company_folder, f"HSN_Summary_{quarter_tag}.xlsx")

    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
        order, totals = _read_existing_hsn_totals(ws)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = f"HSN {quarter_label}"[:31]
        order, totals = [], {}

    for g in group_items_by_hsn(items_list):
        hsn = g["hsn"]
        if hsn not in totals:
            totals[hsn] = {"qty": 0, "amount": 0.0, "cgst": 0.0, "sgst": 0.0, "igst": 0.0}
            order.append(hsn)
        t = totals[hsn]
        t["qty"] += g["qty"]
        t["amount"] += g["amount"]
        t["cgst"] += g["cgst"]
        t["sgst"] += g["sgst"]
        t["igst"] += g["igst"]

    _write_hsn_summary_sheet(ws, order, totals)

    safe_file_save(lambda: wb.save(file_path), file_path)
    return file_path

